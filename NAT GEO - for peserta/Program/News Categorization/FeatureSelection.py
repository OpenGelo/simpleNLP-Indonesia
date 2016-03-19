import openpyxl
import xlrd
import math

__author__ = 'dee'

def mutualInformation():
    fileFeatureMI = openpyxl.Workbook()
    dataFeatureMI = fileFeatureMI.active

    filePreprocess = xlrd.open_workbook('dataset_preprocessing.xlsx') # ambil data label
    dataPreprocess = filePreprocess.sheet_by_index(0)
    rowLenPreprocess = dataPreprocess.nrows

    totalToken = 0
    totalLabel = 1
    label = []

    for i in range(0, rowLenPreprocess):
        data = dataPreprocess.cell(i, 0).value
        tokens = data.split()

        if (i == 0):
            label.append(dataPreprocess.cell(i, 1).value)

        if (i>0 and (dataPreprocess.cell(i, 1).value != dataPreprocess.cell(i-1, 1).value)):
            totalLabel += 1
            label.append(dataPreprocess.cell(i, 1).value)

        for j in range(0, len(tokens)):
            totalToken += 1

    tokenOnLabel = []

    for i in range(0, len(label)):
        for j in range(0, rowLenPreprocess):
            if(dataPreprocess.cell(j, 1).value == label[i]):
                data = dataPreprocess.cell(j, 0).value
                tokens = data.split()
                for k in range(0, len(tokens)):
                    tokenOnLabel.append([])
                    tokenOnLabel[i].append(tokens[k])

    dataFeatureMI.append(['TOKEN', 'LABEL', 'N11', 'N01', 'N10', 'N00', 'I(U;C)'])
    for i in range(0, len(label)):
        data = set(tokenOnLabel[i])
        data = list(data)

        for j in range(0, len(data)):

            n10 = 0
            n00 = 0
            for k in range(0, len(label)):
                if (label[i] != label[k]):
                    n10 += tokenOnLabel[k].count(data[j])

            # for k in range(0, len(label)):
            #     if (label[i] != label[k]):
            #
            #             n00 += tokenOnLabel[k].count(not (data[j]))


            n11 = tokenOnLabel[i].count(data[j])
            n01 = len(tokenOnLabel[i]) - n11
            n00 = totalToken - n11 - n10 - n01

            a = ((n11/totalToken) * math.log((totalToken*n11 / ((n11+n10)*(n11+n01)) ), 2))
            b = ((n01/totalToken) * math.log((totalToken*n01 / ((n01+n00)*(n01+n11)) ), 2))
            d = ((n00/totalToken) * math.log((totalToken*n00 / ((n00+n01)*(n00+n10)) ), 2))

            if((n10) == 0):
                c = 0
                iuc = a + b - c + d
                # iuc = math.log( ((n11 * totalToken) / (n11 + n10) * (n11 + n01)), 2 )
            else:
                c = ((n10/totalToken) * math.log((totalToken*n10) / ((n10+n11)*(n10+n00)), 2))
                iuc = a + b - c + d
            # iuc = math.log( ((n11 * totalToken) / ((n11 + n10) * (n11 + n01))), 2 )

            # print n11, n10, n01, n00, totalToken, iuc
            # print a, b, c, d

            dataFeatureMI.append([data[j], label[i], n11, n01, n10, n00, iuc])
    fileFeatureMI.save("dataset_MIvalue.xlsx") # save hasil feature


def elimination():
    fileFeature = openpyxl.Workbook()
    dataFeature = fileFeature.active

    fileFeatureMI = xlrd.open_workbook('dataset_MIvalue.xlsx') # ambil data hasil mutual information
    dataFeatureMI = fileFeatureMI.sheet_by_index(0)
    rowLenFeatureMI = dataFeatureMI.nrows-1

    label = []
    term = []
    n = []
    avg = []

    for i in range(0, rowLenFeatureMI):
        label.append(dataFeatureMI.cell(i+1, 1).value)
        term.append(dataFeatureMI.cell(i+1, 0).value)
    labelSet = set(label)
    labelSet = list(labelSet)
    termSet = set(term)
    termSet = list(termSet)

    #calc average avg per class
    for i in range(0, len(labelSet)):
        n.append(0)
        totalMIvalue = 0
        for j in range(0, rowLenFeatureMI):
            if (dataFeatureMI.cell(j+1, 1).value == labelSet[i]):
                n[i] += 1
                totalMIvalue += dataFeatureMI.cell(j+1, 6).value
        avg.append(totalMIvalue / n[i])
        print (labelSet[i], totalMIvalue, n[i], avg[i])

    dataFeature.append(['TOKEN', 'LABEL', 'N11', 'N01', 'N10', 'N00', 'I(U;C)', 'AVG'])
    k = 2
    for i in range(0, len(labelSet)):
        for j in range(0, rowLenFeatureMI):
            if ((dataFeatureMI.cell(j+1, 1).value == labelSet[i]) and (dataFeatureMI.cell(j+1, 6).value >= avg[i])):
                dataFeature.cell(row=k, column=1).value = dataFeatureMI.cell(j+1, 0).value
                dataFeature.cell(row=k, column=2).value = dataFeatureMI.cell(j+1, 1).value
                dataFeature.cell(row=k, column=3).value = dataFeatureMI.cell(j+1, 2).value
                dataFeature.cell(row=k, column=4).value = dataFeatureMI.cell(j+1, 3).value
                dataFeature.cell(row=k, column=5).value = dataFeatureMI.cell(j+1, 4).value
                dataFeature.cell(row=k, column=6).value = dataFeatureMI.cell(j+1, 5).value
                dataFeature.cell(row=k, column=7).value = dataFeatureMI.cell(j+1, 6).value
                dataFeature.cell(row=k, column=8).value = avg[i]
                k += 1
    fileFeature.save('dataset_selectedFeature.xlsx') # save hasil elimination