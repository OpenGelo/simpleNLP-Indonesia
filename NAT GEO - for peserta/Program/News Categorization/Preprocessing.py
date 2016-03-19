import re

import xlrd
import openpyxl

__author__ = 'undeed'


def preprocessFile (filenameInput, filenamePreprocessed):
    "fungsi memproses file excel filenameInput, " \
    "dan menyimpan hasil preprocessing ke dalam file excel filenamePreprocessed"

    # buka file input
    fileTrain = xlrd.open_workbook(filenameInput)
    dataTrain = fileTrain.sheet_by_index(0)
    rowLen = dataTrain.nrows

    # siapkan file output
    filePreprocessed = openpyxl.Workbook()
    dataPreprocessed = filePreprocessed.active

    # untuk setiap data input, lakukan preprocessing
    # dan hasil preprocessing simpan ke dalam data output
    for i in range(rowLen):
        data_i = dataTrain.cell(i,0).value
        class_i = dataTrain.cell(i, 1).value
        prep = preprocess(data_i)

        if prep:
            for i in range(len(prep)):
                dataPreprocessed.append([''.join(prep[i]), class_i])

    # simpan file output
    filePreprocessed.save(filenamePreprocessed)
    return dataPreprocessed


def preprocess (article):
    "fungsi melakukan preprocessing sebuah data artikel " \
    "(normalisasi, stemming, tokenisasi) " \
    "dan mengembalikan list token hasil preprocessing"

    # case folding
    article = article.lower()

    # url removal
    article = removeUrl(article)

    # non alphabetic removal
    article = re.sub(r'([a-z])-([a-z])', r'\1 \2', article)
    article = re.sub(r'[^a-z|^ ]', '', article)

    # tokenization
    tokens = article.split()

    # coming soon - stemming
    rootword = [line.rstrip('\n') for line in open('dictionary/rootword.txt')]

    nonKataDasar = list(set(tokens) - set(rootword))
    kataDasar = list(set(tokens) - set(nonKataDasar))

    for i in range(0, len(nonKataDasar)):
        nonKataDasar[i] = stemming(nonKataDasar[i],rootword)

    tokens = list(set(nonKataDasar + kataDasar))

    # stopword removal
    result = stopwordRemoval(tokens)
    return result


def removeUrl(article):
    "fungsi menghapus URL di dalam sebuah artikel"

    def regex_or(*items):
        r = '|'.join(items)
        r = '(' + r + ')'
        return r

    def pos_lookahead(r):
        return '(?=' + r + ')'

    def neg_lookahead(r):
        return '(?!' + r + ')'

    def optional(r):
        return '(%s)?' % r

    PunctChars = r'''['".?!,:;]'''
    Punct = '%s+' % PunctChars
    Entity = '&(amp|lt|gt|quot);'

    UrlStart1 = regex_or(r'https?://?', r'www\.')
    CommonTLDs = regex_or('com', 'co\\.uk', 'org', 'net', 'info', 'ca')
    UrlStart2 = r'[a-z0-9\.-]+?' + r'\.' + CommonTLDs + pos_lookahead(r'[/ \W\b]')
    UrlBody = r'[^ \t\r\n<>]*?'
    UrlExtraCrapBeforeEnd = '%s+?' % regex_or(PunctChars, Entity)
    UrlEnd = regex_or(r'\.\.+', r'[<>]', r'\s', '$')
    Url = (r'\b' +
           regex_or(UrlStart1, UrlStart2) +
           UrlBody +
           pos_lookahead(optional(UrlExtraCrapBeforeEnd) + UrlEnd))

    Url_RE = re.compile("(%s)" % Url, re.U | re.I)
    article = re.sub(Url_RE, "", article)

    return article


def stopwordRemoval(tokens):
    "proses menghapus token-token yang tidak bermakna " \
    "berdasarkan daftar stopword"

    stopword = [line.rstrip('\n') for line in open('dictionary/stopword.txt')]
    stopword2 = [line.rstrip('\n') for line in open('dictionary/dict_noise.txt')]
    tokens2 = []
    for ii in range(0, len(tokens)):
        if tokens[ii] not in stopword and tokens[ii] not in stopword2:
            tokens2.append(tokens[ii])
    return tokens2


def firstRule(token):
    "fungsi stemming menghapus partikel -kah, -lah, -tah, -pun"

    # affix -kah
    if re.search('([a-z0-9]+)kah$',token):
        token = re.sub(r'([a-z0-9]+)kah$', r'\1', token)
        return token

    # affix -lah
    if re.search('([a-z0-9]+)lah$',token):
        token = re.sub(r'([a-z0-9]+)lah$', r'\1', token)
        return token

    # affix -tah
    if re.search('([a-z0-9]+)tah$',token):
        token = re.sub(r'([a-z0-9]+)tah$', r'\1', token)
        return token

    # affix -pun
    if re.search('([a-z0-9]+)pun$',token):
        token = re.sub(r'([a-z0-9]+)pun$', r'\1', token)
        return token
    return token


def secondRule(token):
    "fungsi stemming menghapus kata ganti milik -ku, -mu, -nya"

    # affix -nya
    if re.search('([a-z0-9]+)nya$',token):
        token = re.sub(r'([a-z0-9]+)nya$', r'\1', token)
        return token

    # affix -ku
    if re.search('([a-z0-9]+)ku$',token):
        token = re.sub(r'([a-z0-9]+)ku$', r'\1', token)
        return token

    # affix -mu
    if re.search('([a-z0-9]+)mu$',token):
        token = re.sub(r'([a-z0-9]+)mu$', r'\1', token)
        return token
    return token


def thirdRule(token):
    "fungsi stemming menghapus awalan me-, pe-, di-, ter-, ke-"

    # affix meng-
    if re.search('^meng',token):
        token = re.sub('^meng', '', token)
        return token

    # affix meny-
    if re.search('^meny',token):
        token = re.sub('^meny', 's', token)
        return token

    # affix men-
    if re.search('^men',token):
        token = re.sub('^men', '', token)
        return token

    # affix me-
    if re.search('^me',token):
        token = re.sub('^me', '', token)
        if re.search('^m([^aiueo])',token):
            token = re.sub('^m([^aiueo])',r'\1',token)
        return token

    # # affix mem-
    # if re.search('^mem',token):
    #     if re.search('^mem[aiueo]',token):
    #         token = re.sub('^mem', 'p', token)
    #     else:
    #         token = re.sub('^mem', '', token)
    #     return token
    #
    # # affix me-
    # if re.search('^me',token):
    #     token = re.sub('^me', '', token)
    #     return token

    # affix peng-
    if re.search('^peng',token):
        token = re.sub('^peng', '', token)
        return token

    # affix peny-
    if re.search('^peny',token):
        token = re.sub('^peny', 's', token)
        return token

    # affix pen-
    if re.search('^pen',token):
        token = re.sub('^pen', '', token)
        return token

    # affix pem-
    if re.search('^pem',token):
        if re.search('^pem[aiueo]',token):
            token = re.sub('^pem', 'p', token)
        else:
            token = re.sub('^pem', '', token)
        return token

    # affix di-
    if re.search('^di',token):
        token = re.sub('^di', '', token)
        return token

    # affix ter-
    if re.search('^ter',token):
        token = re.sub('^ter', '', token)
        return token

    # affix ke-
    if re.search('^ke',token):
        token = re.sub('^ke', '', token)
        return token

    return token


def fourthRule(token):
    "fungsi stemming menghapus awalan be-, pe-"

    # affix ber-
    if re.search('^ber',token):
        token = re.sub('^ber', '', token)
        return token

    # affix belajar
    if re.search('^belajar',token):
        token = re.sub('^belajar', 'ajar', token)
        return token

    # affix bek+er-
    if re.search('^bek+er',token):
        token = re.sub('^bek+er', 'ker', token)
        return token

    # affix per-
    if re.search('^per',token):
        token = re.sub('^per', '', token)
        return token

    # affix pelajar
    if re.search('^pelajar',token):
        token = re.sub('^pelajar', 'ajar', token)
        return token

    # affix pe-
    if re.search('^pe',token):
        token = re.sub('^pe', '', token)
        return token

    return token


def fifthRule(token):
    "fungsi stemming menghapus akhiran -kan, -an, -i"

    # affix -kan
    if re.search(r'([a-z0-9]+)kan$',token):
        token = re.sub(r'([a-z0-9]+)kan$', r'\1', token)
        return token

    # affix -an
    if re.search(r'([a-z0-9]+)an$',token):
        token = re.sub(r'([a-z0-9]+)an$', r'\1', token)
        return token

    # affix -i
    if re.search(r'([a-z0-9]+)i$',token):
        token = re.sub(r'([a-z0-9]+)i$', r'\1', token)
        return token

    return token


def stemming(token, rootword):
    "fungsi stemming menggunakan algoritma Porter untuk Bahasa Indonesia"

    # porter stemmer
    # 1. Menghapus partikel seperti: -kah, -lah, -tah, -pun
    # 2. Mengapus kata ganti (Possesive Pronoun), seperti -ku, -mu, -nya
    # 3. Mengapus awalan pertama. Jika tidak ditemukan, maka lanjut ke langkah 4a, dan jika ada  maka lanjut ke langkah 4b.
    # 4 .a. Menghapus awalan kedua, dan dilanjutkan pada langkah ke 5a
    #    b. Menghapus akhiran, jika tidak ditemukan maka kata tersebut diasumsikan sebagai kata  dasar (root word). Jika ditemukan maka lanjut ke langkah 5b.
    # 5. a. Menghapus akhiran dan kata akhir diasumsikan sebagai kata dasar (root word.
    #    b. Menghapus awalan kedua dan kata akhir diasumsikan sebagai kata dasar (root word).


    # 1. ---------------------------
    token = firstRule(token)
    if token in rootword:
        return token

    # 2. ---------------------------
    token = secondRule(token)
    if token in rootword:
        return token

    # 3. ---------------------------
    tempToken = token
    token = thirdRule(token)
    if token in rootword:
        return token
    if tempToken==token:
        token = fourthRule(token)
        if token in rootword:
            return token
        token = fifthRule(token)
        if token in rootword:
            return token
    else:
        temptoken2 = token
        token = fifthRule(token)
        if token in rootword :
            return token
        if token != temptoken2:
            token = fourthRule(token)
            if token in rootword :
                return token
    # 4. ---------------------------
    # 5. ---------------------------
    return token


