from __future__ import division
import Preprocessing
import NaiveBayes
import xlrd
import openpyxl
import FeatureSelection

__author__ = 'undeed'


def train(filename):

    fileTrain = xlrd.open_workbook(filename)
    dataTrain = fileTrain.sheet_by_index(0)
    rowLen = dataTrain.nrows

    filePreprocessed = openpyxl.Workbook()
    dataPreprocessed = filePreprocessed.active

    for i in range(0, rowLen):
        data_i = dataTrain.cell(i,0).value
        class_i = dataTrain.cell(i, 1).value
        prep = Preprocessing.preprocess(data_i)
        # print(prep)

        if prep:
            # prep = list(prep).split()
            for i in range(0,len(prep)):
                dataPreprocessed.append([''.join(prep[i]), class_i])

    filePreprocessed.save("dataset_preprocessing.xlsx")

    # FeatureSelection.mutualInformation()
    # FeatureSelection.elimination()
    # NaiveBayes.classify(filename)