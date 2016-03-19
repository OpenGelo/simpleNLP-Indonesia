from __future__ import division
from nltk.metrics import ConfusionMatrix
import Preprocessing
import openpyxl
import xlrd

__author__ = 'undeed'


class NaiveBayes:

    def __init__(self, fileModel):
        self.fileModel = fileModel

    def learning(self, inputFile, featureFile):
        "fungsi learning untuk membuat model multinomial naive bayes classifier"

        # create file baru untuk result dari model naive bayes
        fileResult = openpyxl.Workbook()
        dataResult = fileResult.active

        # open data training specify by filename
        fileTrain = xlrd.open_workbook(inputFile)
        dataTrain = fileTrain.sheet_by_index(0)
        totalNumDocuments = dataTrain.nrows

        # open data feature
        fileFeature = xlrd.open_workbook(featureFile)
        dataFeature = fileFeature.sheet_by_index(0)
        rowLenFeature = dataFeature.nrows

        label = []
        term = []
        cls = []

        # get term feature and label class in cell dataFeature
        for i in range(0, rowLenFeature):
            term.append(dataFeature.cell(i, 0).value)
            label.append(dataFeature.cell(i, 1).value)

        # ambil data yg unik pada term dan label (tidak ada yg redundan)
        labelSet = list(set(label))
        termSet = list(set(term))

        # set label and pc in xls
        dataResult.append(['Data Probabilitas'])
        dataResult.append(['Prob Class'])

        # get p(c)(probability of class) atau (prior of class) terhadap persebaran class pada data train yang ada
        for i in range(0, len(labelSet)):
            dataResult.cell(row=1, column=i + 2).value = labelSet[i]

            cls.append([])
            cls[i].append(labelSet[i])
            cls[i].append(label.count(labelSet[i]))
            pCj = cls[i][1] / totalNumDocuments  # probabilitas of class

            dataResult.cell(row=2, column=i + 2).value = pCj # masukin ke file baru

        # get p(t|c) (probability of term given class) ata (likelihood of term given class)
        # di loop untuk setiap termset di cari likelihood terhadap semua class yang ada
        for i in range(0, len(termSet)): #loop untuk setiap term
            dataResult.append([termSet[i]]) # masukin dulu datanya ke file baru

            for j in range(0, len(labelSet)): #loop untuk setiap term terhadap setiap class yg ada
                countTerm = 0
                countTermInClass = 0
                idx = 1
                endLoop = False

                while endLoop == False: #mencari term(i) dengan label(j) pada cell dataFeature

                    if dataFeature.cell(idx, 1).value == labelSet[j]:
                        if dataFeature.cell(idx, 0).value == termSet[i]:
                            countTerm = countTerm + 1 #ngitung berapa banyak term(i) pada label(j)

                        countTermInClass = countTermInClass + 1 #ngitung berapa banyak total term pada label(j)
                        idx = idx + 1

                        if idx < rowLenFeature:
                            if dataFeature.cell(idx, 1).value != labelSet[j]: # kalau bukan udah ganti belum labelnya, kalau udah stop aja soalnya datanya terurut by label
                                endLoop = True
                        else:
                            endLoop = True
                    else:
                        idx = idx + 1

                pTC = (countTerm + 1) / ((countTermInClass) + len(termSet)) #ngitung likelihood pTC make smoothing laplace #asli pTC nya mah (countTerm / countTermInClass)
                dataResult.cell(row=3+i, column=j+2).value = pTC

        fileResult.save(self.fileModel)
        return

    def classify(self,article, dataModel):
        "fungsi mengklasifikasikan suatu artikel berdasarkan model Bayes yang telah dibuat"

        label = []

        # preprocessing for data input
        token = Preprocessing.preprocess(article)

        # get label from xls
        for i in range(1, dataModel.ncols):
            label.append(dataModel.cell(0, i).value)

        if token:
            probability = []
            # get probability
            i = 0
            idx = 0
            while i < len(token):
                probability.append([])
                j = 2
                while j < dataModel.nrows:
                    if token[i] == dataModel.cell(j, 0).value:
                        for k in range(0, len(label)):
                            pd = dataModel.cell(j, k+1).value
                            probability[idx].append(pd)
                        break
                    j += 1
                if j == dataModel.nrows:
                    del probability[-1]
                    i += 1
                else:
                    i += 1
                    idx += 1

            # probability calc
            pFinal = []
            for i in range(0, len(label)):
                pc = dataModel.cell(1, i+1).value
                valP = 1
                for j in range(0, len(probability)):
                    valP *= probability[j][i]

                if len(probability) == 1:
                    if valP != 1:
                        value = valP
                    else:
                        value = 0
                else:
                    if valP != 1:
                        value = valP*pc
                    else:
                        value = 0

                pFinal.append(value)

            maks = max(pFinal)

            if maks != 0:
                for i in range(0, len(pFinal)):
                    if pFinal[i] == maks:
                        idxMax = i
                decision = dataModel.cell(0, idxMax+1).value
            else:
                decision = 'ERROR'
        else:
            decision = 'ERROR'

        return decision

    def testing(self, inputFile, outputFile):
        "fungsi testing menguji "

        fileModel = xlrd.open_workbook(self.fileModel)
        dataModel = fileModel.sheet_by_index(0)

        fileResult = openpyxl.Workbook()
        dataResult = fileResult.active

        fileTest = xlrd.open_workbook(inputFile)
        dataTest = fileTest.sheet_by_index(0)
        dataTestLen = dataTest.nrows-1

        decisionTrue = 0
        decisionError = 0
        allCategory = []
        allDec = []

        for i in range(1, dataTestLen+1):
            # print (i)
            article = dataTest.cell(i, 0).value
            decision_actually = dataTest.cell(i, 1).value
            decision_test = self.classify(article, dataModel)

            dataResult.append([article, decision_actually, decision_test])

            if decision_test == 'ERROR':
                decisionError += 1
            elif decision_actually == decision_test:
                decisionTrue += 1

            allCategory.append(decision_actually)
            allDec.append(decision_test)

        confisionMatrix = ConfusionMatrix(allCategory, allDec)
        print (confisionMatrix)

        accuracy = (decisionTrue / (dataTestLen - decisionError)) * 100
        dataResult.append([' ', 'DECISION TRUE', 'DECISION FALSE', 'DECISION ERROR', 'ACCURACY'])
        dataResult.append([' ', decisionTrue, (dataTestLen - decisionError - decisionTrue), decisionError, accuracy])
        fileResult.save(outputFile)
        return