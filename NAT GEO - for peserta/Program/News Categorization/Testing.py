from __future__ import division
from nltk.metrics import ConfusionMatrix
import Preprocessing
import xlrd
import openpyxl

__author__ = 'undeed'


def testOneData(article):

    # load xls classification
    fileClass = xlrd.open_workbook("data_classification.xlsx")
    dataClass = fileClass.sheet_by_index(0)

    label = []
    # preprocessing for data input
    token = Preprocessing.preprocess(article)
    # print(token)

    # get label from xls
    for i in range(1, dataClass.ncols):
        label.append(dataClass.cell(0, i).value)

    if token:
        probability = []
        # get probability
        i = 0
        idx = 0
        while i < len(token):
            probability.append([])
            j = 2
            while j < dataClass.nrows:
                if token[i] == dataClass.cell(j, 0).value:
                    for k in range(0, len(label)):
                        pd = dataClass.cell(j, k+1).value
                        probability[idx].append(pd)
                    break
                j += 1
            if j == dataClass.nrows:
                del probability[-1]
                i += 1
            else:
                i += 1
                idx += 1

        # probability calc
        pFinal = []
        for i in range(0, len(label)):
            pc = dataClass.cell(1, i+1).value
            valP = 1
            for j in range(0, len(probability)):
                valP *= probability[j][i]

            if len(probability) == 1:
                if valP != 1:
                    value = valP
                else:
                    value = 0
            else:
                if valP != 1:
                    value = valP*pc
                else:
                    value = 0

            pFinal.append(value)

        maks = max(pFinal)

        if maks != 0:
            for i in range(0, len(pFinal)):
                if pFinal[i] == maks:
                    idxMax = i
            decision = dataClass.cell(0, idxMax+1).value
        else:
            decision = 'ERROR'
    else:
        decision = 'ERROR'

    return decision


def testOneFile(fileName):
    fileResult = openpyxl.Workbook()
    dataResult = fileResult.active

    fileTest = xlrd.open_workbook(fileName)
    dataTest = fileTest.sheet_by_index(0)
    dataTestLen = dataTest.nrows-1

    decisionTrue = 0
    decisionError = 0
    allCategory = []
    allDec = []

    for i in range(1, dataTestLen+1):
        # print (i)
        article = dataTest.cell(i, 0).value
        decision_actually = dataTest.cell(i, 1).value
        decision_test = testOneData(article)

        # print decision_actually, decision_test

        # print decision_actually, '----------', decision_test
        dataResult.append([article, decision_actually, decision_test])

        if decision_test == 'ERROR':
            decisionError += 1
        elif decision_actually == decision_test:
            decisionTrue += 1

        allCategory.append(decision_actually)
        allDec.append(decision_test)

    confisionMatrix = ConfusionMatrix(allCategory, allDec)
    print (confisionMatrix)

    accuracy = (decisionTrue / (dataTestLen - decisionError)) * 100
    dataResult.append([' ', 'DECISION TRUE', 'DECISION FALSE', 'DECISION ERROR', 'ACCURACY'])
    dataResult.append([' ', decisionTrue, (dataTestLen - decisionError - decisionTrue), decisionError, accuracy])
    fileResult.save("RESULT_train-train.xlsx")